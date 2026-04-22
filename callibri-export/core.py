"""
core.py — провайдер-агностик: период, конфиг, запись файлов, оркестрация.

Вся специфика API (Callibri, Calltouch) — в пакете providers/.
run_export определяет провайдера для каждого проекта по ключу "provider"
и делегирует выгрузку.
"""

import csv
import os
import re
import sys
import json
import logging
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import providers

log = logging.getLogger(__name__)


# ── Утилиты ──────────────────────────────────────────────────────────────────

def get_app_dir():
    """Директория приложения (рядом с .exe или рядом с .py)."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def parse_date(value, name="date"):
    """Парсим дату из строки dd.mm.yyyy. Бросает ValueError при ошибке."""
    try:
        return datetime.strptime(value, "%d.%m.%Y")
    except ValueError:
        raise ValueError(
            f"Неверный формат даты {name}='{value}'. Ожидается dd.mm.yyyy (например, 01.03.2026)"
        )


def resolve_period(date1_str=None, date2_str=None, days=None):
    """Определяем период выгрузки. Возвращает (date1, date2)."""
    if date1_str and date2_str:
        date1 = parse_date(date1_str, "date1")
        date2 = parse_date(date2_str, "date2")
        if date1 > date2:
            raise ValueError(
                f"Начальная дата ({date1_str}) позже конечной ({date2_str})"
            )
    elif days:
        if days < 1:
            raise ValueError(f"days должен быть >= 1, получено: {days}")
        date2 = datetime.now()
        date1 = date2 - timedelta(days=days - 1)
    else:
        date2 = datetime.now()
        date1 = date2 - timedelta(days=6)
    return date1, date2


def sanitize_filename(name):
    """Убираем из имени символы, недопустимые в именах файлов."""
    return re.sub(r'[<>:"/\\|?*]', "_", name).strip()


# ── Конфигурация проектов ────────────────────────────────────────────────────

def load_projects(path=None):
    """Загружаем projects.json. Бросает FileNotFoundError / ValueError."""
    if path is None:
        path = os.path.join(get_app_dir(), "projects.json")
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Файл projects.json не найден: {path}\n"
            'Создай его рядом с приложением. Пример: [{"site_id": 4112, "folder": "energocenter"}]'
        )
    with open(path, encoding="utf-8") as f:
        projects_data = json.load(f)
    if not projects_data:
        raise ValueError("projects.json пустой — добавь хотя бы один проект.")
    return projects_data


def save_projects(projects_data, path=None):
    """Сохраняем projects.json."""
    if path is None:
        path = os.path.join(get_app_dir(), "projects.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(projects_data, f, ensure_ascii=False, indent=2)


def get_project_provider(proj_conf):
    """Вернуть модуль провайдера для проекта (по ключу 'provider', дефолт callibri)."""
    return providers.get_provider(proj_conf.get("provider"))


# ── Запись файлов ────────────────────────────────────────────────────────────

# Префиксы, которые Excel/LibreOffice интерпретирует как формулу.
# Значение из API, начинающееся на один из них, потенциально выполнит
# произвольную формулу у получателя файла (CSV/Formula Injection).
_FORMULA_PREFIX = ("=", "+", "-", "@", "\t", "\r")


def _csv_safe(value):
    """Экранировать значение от формула-инъекции в CSV/XLSX."""
    if value is None:
        return ""
    s = str(value)
    if s.startswith(_FORMULA_PREFIX):
        return "'" + s
    return s


def write_csv(filepath, rows, columns):
    """CSV с заголовком (разделитель ;, UTF-8 BOM). Защита от формула-инъекции."""
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(columns)
        for row in rows:
            writer.writerow([_csv_safe(row.get(col, "")) for col in columns])


def write_xlsx(filepath, rows, columns):
    """XLSX с заголовком и автошириной колонок. Защита от формула-инъекции."""
    wb = Workbook()
    ws = wb.active

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([_csv_safe(row.get(col, "")) for col in columns])

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

    wb.save(filepath)


# ── Google Sheets ───────────────────────────────────────────────────────────

def _emit(on_log, message):
    if on_log:
        on_log(message)
    else:
        log.info(message)


def _export_to_gsheet(credentials_path, gsheet_conf, rows, columns, on_log):
    """Отправка данных в Google Sheets. Ошибки не прерывают экспорт."""
    try:
        import gsheets as gs
    except ImportError:
        _emit(on_log, "  Google Sheets: пакеты gspread/google-auth не установлены — пропускаем")
        return

    spreadsheet_id = gsheet_conf.get("spreadsheet_id", "")
    sheet_name = gsheet_conf.get("sheet_name", "")
    mode = gsheet_conf.get("mode", "append")

    if not spreadsheet_id or not sheet_name:
        _emit(on_log, "  Google Sheets: не указана таблица или лист — пропускаем")
        return

    try:
        client = gs.authorize(credentials_path)
        gs.export_to_sheet(
            client, spreadsheet_id, sheet_name,
            rows, columns, mode=mode, on_log=on_log,
        )
    except Exception as e:
        _emit(on_log, f"  Google Sheets: Ошибка — {e}")


# ── Главная функция экспорта ─────────────────────────────────────────────────

def run_export(
    credentials,
    date1_str=None, date2_str=None, days=None,
    projects_path=None,
    output_dir=None,
    enabled_keys=None,
    on_log=None,
    on_progress=None,
    gsheet_credentials=None,
):
    """
    Главная функция экспорта.

    Параметры:
      credentials — dict: {provider_name: {field: value}}
        Пример: {"callibri": {"email": "...", "token": "..."}}
      date1_str, date2_str — период (dd.mm.yyyy), или days для последних N дней
      projects_path — путь к projects.json (None = рядом с приложением)
      output_dir — папка результатов (None = output/ рядом с приложением)
      enabled_keys — set of (provider_name, site_id) tuples; если передан —
        переопределяет "enabled" из конфига
      on_log(msg) — callback логов
      on_progress(project_idx, total_projects, chunk_idx, total_chunks) — прогресс
      gsheet_credentials — путь к credentials.json для Google Sheets

    Бросает ValueError при ошибках валидации.
    Возвращает dict с результатами.
    """
    # 1. Период
    date1, date2 = resolve_period(date1_str, date2_str, days)
    _emit(on_log, f"Период: {date1.strftime('%d.%m.%Y')} — {date2.strftime('%d.%m.%Y')}")

    # 2. Загружаем конфигурацию
    projects_config = load_projects(projects_path)
    _emit(on_log, f"Проектов в конфиге: {len(projects_config)}")

    # 3. Определяем активные проекты и валидируем учётные данные для них
    def _is_enabled(proj):
        provider_name = proj.get("provider", "callibri")
        site_id = proj.get("site_id")
        if enabled_keys is not None:
            return (provider_name, site_id) in enabled_keys
        return proj.get("enabled", True)

    active_projects = [p for p in projects_config if _is_enabled(p)]

    # 4. Кэш list_sites per-provider (чтобы не звать API несколько раз)
    sites_cache = {}   # provider_name -> {site_id: site_dict}

    def _get_sites_lookup(provider):
        if provider.NAME not in sites_cache:
            _emit(on_log, f"Загружаем список проектов из API [{provider.LABEL}]...")
            creds = credentials.get(provider.NAME) or {}
            sites = provider.list_sites(creds)
            sites_cache[provider.NAME] = {s.get("site_id"): s for s in sites}
        return sites_cache[provider.NAME]

    # 5. Выгрузка
    if output_dir is None:
        output_dir = os.path.join(get_app_dir(), "output")
    date_suffix = f"{date1.strftime('%d%m%Y')}-{date2.strftime('%d%m%Y')}"

    processed = 0
    disabled = 0
    errors = 0
    total_failed_chunks = 0
    report = []
    total_active = len(active_projects)
    project_idx_among_active = 0

    for proj_conf in projects_config:
        provider_name = proj_conf.get("provider", "callibri")
        site_id = proj_conf.get("site_id")
        folder = proj_conf.get("folder")

        if not _is_enabled(proj_conf):
            disabled += 1
            continue

        try:
            provider = providers.get_provider(provider_name)
        except ValueError as e:
            _emit(on_log, f"ОШИБКА: {e} (site_id={site_id}, folder={folder})")
            errors += 1
            continue

        creds = credentials.get(provider.NAME) or {}
        ok, msg = provider.check_credentials(creds)
        if not ok:
            _emit(on_log, f"ОШИБКА [{provider.LABEL}]: {msg} (folder={folder})")
            errors += 1
            continue

        manual_only = getattr(provider, "REQUIRES_MANUAL_SITE_ID", False)
        if manual_only:
            site = {"site_id": site_id, "sitename": proj_conf.get("name") or folder or str(site_id)}
        else:
            try:
                sites_lookup = _get_sites_lookup(provider)
            except Exception as e:
                _emit(on_log, f"ОШИБКА загрузки списка [{provider.LABEL}]: {e}")
                errors += 1
                continue

            if site_id not in sites_lookup:
                _emit(on_log, f"ПРЕДУПРЕЖДЕНИЕ: site_id={site_id} (folder={folder}) не найден в {provider.LABEL}")
                errors += 1
                continue

            site = sites_lookup[site_id]

        site_name = site.get("sitename") or site.get("name") or str(site_id)

        channel_filter = proj_conf.get("channels")
        columns = proj_conf.get("fields") or list(provider.DEFAULT_COLUMNS)
        type_filter = proj_conf.get("types")
        status_filter = proj_conf.get("statuses")
        split_by_channel = proj_conf.get("split_by_channel", False)
        out_format = proj_conf.get("format", "xlsx")
        file_export = proj_conf.get("file_export", True)

        _emit(on_log, f"[{provider.LABEL}] {site_name} (id={site_id}) → output/{folder}/")
        if channel_filter:
            _emit(on_log, f"  Фильтр каналов: {channel_filter}")
        if type_filter:
            _emit(on_log, f"  Фильтр типов: {type_filter}")
        if status_filter:
            _emit(on_log, f"  Фильтр статусов: {status_filter}")
        if columns != list(provider.DEFAULT_COLUMNS):
            _emit(on_log, f"  Поля: {columns}")

        # Callback прогресса по чанкам
        _pi = project_idx_among_active

        def _on_chunk(cur_chunk, tot_chunks, _pi=_pi):
            if on_progress:
                on_progress(_pi, total_active, cur_chunk, tot_chunks)

        filters = {
            "channels": channel_filter,
            "columns": columns,
            "types": type_filter,
            "statuses": status_filter,
        }

        try:
            has_data, rows_by_channel, failed_chunks = provider.process_site(
                site, date1, date2, creds, filters, on_log, _on_chunk
            )
        except Exception as e:
            _emit(on_log, f"ОШИБКА при выгрузке [{provider.LABEL}] {folder}: {e}")
            errors += 1
            project_idx_among_active += 1
            continue

        total_failed_chunks += failed_chunks

        if not has_data:
            errors += 1
            report.append((site_name, 0))
            project_idx_among_active += 1
            continue

        all_rows = [row for rows in rows_by_channel.values() for row in rows]
        all_rows.sort(key=lambda r: r.get("date", ""))
        project_total = len(all_rows)

        # Файл
        if file_export:
            project_dir = os.path.join(output_dir, folder)
            os.makedirs(project_dir, exist_ok=True)

            write_fn = write_csv if out_format == "csv" else write_xlsx
            ext = out_format
            prefix = f"{provider.NAME}_{date_suffix}"

            if split_by_channel:
                for ch_name, rows in rows_by_channel.items():
                    rows.sort(key=lambda r: r.get("date", ""))
                    safe_name = sanitize_filename(ch_name)
                    filepath = os.path.join(project_dir, f"{prefix}_{safe_name}.{ext}")
                    write_fn(filepath, rows, columns)
                    _emit(on_log, f"  [{ch_name}] → {len(rows)} строк → {os.path.basename(filepath)}")
            else:
                filepath = os.path.join(project_dir, f"{prefix}.{ext}")
                write_fn(filepath, all_rows, columns)
                _emit(on_log, f"  Сохранено: {project_total} строк → {os.path.basename(filepath)}")
        else:
            _emit(on_log, f"  Файловый экспорт выключен — {project_total} строк")

        if failed_chunks:
            _emit(on_log, f"  Внимание: {failed_chunks} чанк(ов) не удалось загрузить — данные неполные")

        # Google Sheets
        gsheet_conf = proj_conf.get("gsheet")
        if (gsheet_conf and gsheet_conf.get("enabled")
                and gsheet_credentials and all_rows):
            _export_to_gsheet(
                gsheet_credentials, gsheet_conf, all_rows, columns, on_log
            )

        report.append((site_name, project_total))
        processed += 1
        project_idx_among_active += 1

    result = {
        "processed": processed,
        "disabled": disabled,
        "errors": errors,
        "failed_chunks": total_failed_chunks,
        "report": report,
    }

    summary = f"Готово! Обработано: {processed}"
    if report:
        total_rows = sum(c for _, c in report)
        summary += f", строк: {total_rows}"
    _emit(on_log, summary)

    return result
